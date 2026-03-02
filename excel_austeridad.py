# ============================================================================
# GENERADOR DE EXCEL - DASHBOARD DE AUSTERIDAD
# ============================================================================

import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from config import (
    PARTIDAS_AUSTERIDAD, DENOMINACIONES_AUSTERIDAD,
    formatear_fecha, obtener_ultimo_dia_habil
)


def generar_excel_austeridad(datos_dashboard, ur_codigo, ur_nombre, año_anterior=2024, año_actual=2025):
    """
    Genera el archivo Excel del Dashboard de Austeridad.
    
    Args:
        datos_dashboard: Lista de dicts con datos por partida
        ur_codigo: Código de la UR (ej: '100')
        ur_nombre: Nombre de la UR (ej: 'Secretaría')
        año_anterior: Año del ejercido anterior
        año_actual: Año actual
    
    Returns:
        bytes: Contenido del archivo Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard Austeridad"
    
    # =========================================================================
    # ESTILOS
    # =========================================================================
    
    # Colores institucionales SADER
    fill_vino = PatternFill(start_color='722F37', end_color='722F37', fill_type='solid')
    fill_beige = PatternFill(start_color='E6D194', end_color='E6D194', fill_type='solid')
    fill_gray = PatternFill(start_color='D9D9D6', end_color='D9D9D6', fill_type='solid')
    fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    # Fuentes
    font_header = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    font_title = Font(name='Calibri', size=12, bold=True)
    font_subtitle = Font(name='Calibri', size=11, bold=True)
    font_data = Font(name='Calibri', size=10)
    font_notes = Font(name='Calibri', size=9)
    
    # Bordes
    border_thin = Border(
        top=Side(style='thin'),
        bottom=Side(style='thin'),
        left=Side(style='thin'),
        right=Side(style='thin')
    )
    border_none = Border()
    
    # Alineaciones
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_right = Alignment(horizontal='right', vertical='center')
    
    # Formatos numéricos
    fmt_money = '#,##0.00'
    fmt_pct = '0.00%'
    
    # =========================================================================
    # ANCHOS DE COLUMNA
    # =========================================================================
    
    anchos = {
        'A': 10,      # Partida
        'B': 70,      # Denominación
        'C': 18,      # Ejercido año anterior
        'D': 14,      # Original
        'E': 14,      # Modificado
        'F': 14,      # Ejercido Real
        'G': 18,      # Solicitud de pago
        'H': 60,      # Nota
        'I': 14,      # Avance anual
    }
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho
    
    # =========================================================================
    # ENCABEZADOS
    # =========================================================================
    
    # Obtener fecha actual
    hoy = date.today()
    ultimo_habil = obtener_ultimo_dia_habil(hoy)
    
    # Fila 1: Título con fecha
    ws.merge_cells('A1:I1')
    titulo = f'Estado del ejercicio del 1 de enero al {ultimo_habil.day} de {["enero","febrero","marzo","abril","mayo","junio","julio","agosto","septiembre","octubre","noviembre","diciembre"][ultimo_habil.month-1]} de {año_actual} de partidas sujetas a Austeridad'
    ws['A1'] = titulo
    ws['A1'].font = font_title
    ws['A1'].alignment = align_center
    ws.row_dimensions[1].height = 25
    
    # Fila 2: Subtítulo UR
    ws.merge_cells('A2:I2')
    ws['A2'] = f'{ur_codigo}.- {ur_nombre}'
    ws['A2'].font = font_subtitle
    ws['A2'].alignment = align_center
    ws.row_dimensions[2].height = 20
    
    # Fila 3: vacía
    ws.row_dimensions[3].height = 10
    
    # Fila 4: Título de sección
    ws.merge_cells('A4:I4')
    ws['A4'] = 'Partidas sujetas a Austeridad Republicana'
    ws['A4'].font = font_subtitle
    ws['A4'].alignment = align_left
    ws.row_dimensions[4].height = 20
    
    # Fila 5: Primera fila de encabezados
    # Partida, Denominación, Ejercicio fiscal (merge C-G), Nota, Avance anual
    ws['A5'] = 'Partida'
    ws['B5'] = 'Denominación'
    ws.merge_cells('C5:G5')
    ws['C5'] = 'Ejercicio fiscal'
    ws['H5'] = 'Nota'
    ws['I5'] = 'Avance anual'
    
    for col in ['A', 'B', 'C', 'H', 'I']:
        ws[f'{col}5'].font = font_header
        ws[f'{col}5'].fill = fill_vino
        ws[f'{col}5'].alignment = align_center
        ws[f'{col}5'].border = border_thin
    # Aplicar estilo a celdas merged
    for col in ['D', 'E', 'F', 'G']:
        ws[f'{col}5'].fill = fill_vino
        ws[f'{col}5'].border = border_thin
    ws.row_dimensions[5].height = 20
    
    # Fila 6: Segunda fila de encabezados
    # Vacío, Vacío, Ejercido en [año_anterior], [año_actual] (merge D-G), vacío, vacío
    ws['A6'] = ''
    ws['B6'] = ''
    ws['C6'] = f'Ejercido en {año_anterior}'
    ws.merge_cells('D6:G6')
    ws['D6'] = str(año_actual)
    ws['H6'] = ''
    ws['I6'] = ''
    
    for col in ['A', 'B', 'C', 'D', 'H', 'I']:
        ws[f'{col}6'].font = font_header
        ws[f'{col}6'].fill = fill_vino
        ws[f'{col}6'].alignment = align_center
        ws[f'{col}6'].border = border_thin
    for col in ['E', 'F', 'G']:
        ws[f'{col}6'].fill = fill_vino
        ws[f'{col}6'].border = border_thin
    ws.row_dimensions[6].height = 20
    
    # Fila 7: Tercera fila de encabezados (subcolumnas del año actual)
    ws['A7'] = ''
    ws['B7'] = ''
    ws['C7'] = ''
    ws['D7'] = 'Original'
    ws['E7'] = 'Modificado'
    ws['F7'] = 'Ejercido Real'
    ws['G7'] = 'Solicitud de pago por parte de las UR'
    ws['H7'] = ''
    ws['I7'] = ''
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws[f'{col}7'].font = font_header
        ws[f'{col}7'].fill = fill_vino
        ws[f'{col}7'].alignment = align_center
        ws[f'{col}7'].border = border_thin
    ws.row_dimensions[7].height = 35
    
    # Merge vertical para encabezados
    ws.merge_cells('A5:A7')
    ws.merge_cells('B5:B7')
    ws.merge_cells('C6:C7')
    ws.merge_cells('H5:H7')
    ws.merge_cells('I5:I7')
    
    # =========================================================================
    # DATOS CON FÓRMULAS
    # =========================================================================
    
    fila = 8
    for dato in datos_dashboard:
        # A: Partida
        ws.cell(row=fila, column=1, value=dato['Partida'])
        ws.cell(row=fila, column=1).font = font_data
        ws.cell(row=fila, column=1).alignment = align_center
        ws.cell(row=fila, column=1).border = border_thin
        
        # B: Denominación
        ws.cell(row=fila, column=2, value=dato['Denominacion'])
        ws.cell(row=fila, column=2).font = font_data
        ws.cell(row=fila, column=2).alignment = align_left
        ws.cell(row=fila, column=2).border = border_thin
        
        # C: Ejercido Anterior
        ws.cell(row=fila, column=3, value=dato['Ejercido_Anterior'])
        ws.cell(row=fila, column=3).font = font_data
        ws.cell(row=fila, column=3).number_format = fmt_money
        ws.cell(row=fila, column=3).alignment = align_right
        ws.cell(row=fila, column=3).border = border_thin
        
        # D: Original
        ws.cell(row=fila, column=4, value=dato['Original'])
        ws.cell(row=fila, column=4).font = font_data
        ws.cell(row=fila, column=4).number_format = fmt_money
        ws.cell(row=fila, column=4).alignment = align_right
        ws.cell(row=fila, column=4).border = border_thin
        
        # E: Modificado
        ws.cell(row=fila, column=5, value=dato['Modificado'])
        ws.cell(row=fila, column=5).font = font_data
        ws.cell(row=fila, column=5).number_format = fmt_money
        ws.cell(row=fila, column=5).alignment = align_right
        ws.cell(row=fila, column=5).border = border_thin
        
        # F: Ejercido Real
        ws.cell(row=fila, column=6, value=dato['Ejercido_Real'])
        ws.cell(row=fila, column=6).font = font_data
        ws.cell(row=fila, column=6).number_format = fmt_money
        ws.cell(row=fila, column=6).alignment = align_right
        ws.cell(row=fila, column=6).border = border_thin
        
        # G: Solicitud de pago (vacío para llenado manual)
        ws.cell(row=fila, column=7, value='')
        ws.cell(row=fila, column=7).font = font_data
        ws.cell(row=fila, column=7).number_format = fmt_money
        ws.cell(row=fila, column=7).alignment = align_right
        ws.cell(row=fila, column=7).border = border_thin
        
        # H: Nota (FÓRMULA)
        # =SI(Y(F>C),"Monto ejercido...",SI(Y(C=0,E>0),"Solicitar...",SI(...)))
        formula_nota = (
            f'=IF(AND(F{fila}>C{fila},C{fila}>0),"Monto ejercido real mayor al presupuesto ejercido en {año_anterior}.",'
            f'IF(AND(C{fila}=0,E{fila}>0),"Solicitar dictamen antes de ejercer recursos en esta partida.",'
            f'IF(AND(C{fila}=0,F{fila}>0),"Monto ejercido real mayor al presupuesto ejercido en {año_anterior}.",'
            f'IF(AND(F{fila}+G{fila}>C{fila},C{fila}>0),"Solicitar dictamen antes de ejercer recursos en esta partida.",'
            f'IF(AND(C{fila}=0,E{fila}=0,F{fila}=0),"",'
            f'IF(AND(E{fila}>C{fila},F{fila}<C{fila}),"Solicitar dictamen antes de sobrepasar el monto ejercido en {año_anterior}.",'
            f'"Sin observaciones."))))))'
        )
        ws.cell(row=fila, column=8, value=formula_nota)
        ws.cell(row=fila, column=8).font = font_notes
        ws.cell(row=fila, column=8).alignment = align_left
        ws.cell(row=fila, column=8).border = border_thin
        
        # I: Avance anual (FÓRMULA)
        # =SI(Y(C=0,(F>0)+G),"Incremento en presupuesto",(SI.ERROR(((F+G)/C),"")))
        formula_avance = (
            f'=IF(AND(C{fila}=0,OR(F{fila}>0,G{fila}>0)),"Incremento",'
            f'IFERROR((F{fila}+G{fila})/C{fila},""))'
        )
        ws.cell(row=fila, column=9, value=formula_avance)
        ws.cell(row=fila, column=9).font = font_data
        ws.cell(row=fila, column=9).number_format = fmt_pct
        ws.cell(row=fila, column=9).alignment = align_center
        ws.cell(row=fila, column=9).border = border_thin
        
        # Altura de fila
        ws.row_dimensions[fila].height = 28
        
        fila += 1
    
    # =========================================================================
    # FUENTE
    # =========================================================================
    
    fila += 1
    ws.merge_cells(f'A{fila}:I{fila}')
    ws[f'A{fila}'] = 'Fuente: SICOP'
    ws[f'A{fila}'].font = Font(name='Calibri', size=9)
    ws[f'A{fila}'].alignment = align_left
    
    # =========================================================================
    # GUARDAR
    # =========================================================================
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()
